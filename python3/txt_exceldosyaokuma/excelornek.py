import openpyxl
workbook = openpyxl.load_workbook("besNbirK.xlsx")
dataframe1 = workbook.active
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
for satir in range(1,7):
    for sutun in range(1,7):
        print("  "+str(dataframe1.cell(satir,sutun).value)+"  ",end="")
    print()

        
        